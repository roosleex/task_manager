# Additional formats for import_export
from import_export.formats.base_formats import Format
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from io import BytesIO
import os
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import Paragraph
from reportlab.lib.pagesizes import A4, landscape
from common.import_export import is_report_resource

from import_export.formats import base_formats
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import math



def clean_cell(value):
    """Convert None or empty-like values to empty string."""
    # return str(value)
    if value is None:
        return ""
    if isinstance(value, str) and value.strip().lower() == "none":
        return ""
    # Extend with more checks if needed, e.g., numpy.nan
    # return str(value)
    return value



def _insert_line_breaks(text, words_per_line=2):
    parts = str(text).split()
    if len(parts) <= words_per_line:
        return text
    lines = []
    for i in range(0, len(parts), words_per_line):
        lines.append(" ".join(parts[i:i+words_per_line]))
    return "\n".join(lines)


 
class PDF(Format):
    """Custom PDF export format for django-import-export"""

    def get_title(self):
        return "pdf"

    def get_extension(self):
        return "pdf"

    def get_content_type(self):
        return "application/pdf"

    def can_export(self):
        return True  # required for admin

    def can_import(self):
        return False

    def export_data(self, dataset, **kwargs):
        # print(f"kwargsss = {kwargs}")
        is_totals_row = False
        resource = kwargs.get("resource")  # instance of the resource
        if resource:
            if is_report_resource(resource):
                is_totals_row = True

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),   # 👈 Landscape orientation
            leftMargin=10,
            rightMargin=10,
            topMargin=10,
            bottomMargin=10
        )

        # Path to TTF font (must support Unicode)
        font_path = os.path.join('static', 'fonts', 'DejaVuSans.ttf')
        pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
        bold_font_path = os.path.join('static', 'fonts', 'DejaVuSans-Bold.ttf')
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_font_path))

        lighter_green = colors.Color(red=0.85, green=1, blue=0.85)

        styles = getSampleStyleSheet()

        body_style = ParagraphStyle(
            "BodyStyle",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=10,
            leading=12
        )
        bold_body_style = ParagraphStyle(
            "BoldBodyStyle",
            parent=body_style,
            fontName="DejaVuSans-Bold",
            fontSize=11,
            leading=12
        )
        header_style = ParagraphStyle(
            "HeaderStyle",
            parent=styles["Heading4"],
            fontName="DejaVuSans",
            fontSize=12,
            leading=14,
            spaceAfter=6
        )

        page_width, page_height = landscape(A4)
        left_margin = right_margin = 36 # 0.5 inch margins
        usable_width = page_width - left_margin - right_margin

        # Wrap headers
        table_data = [[Paragraph(str(h), header_style) for h in dataset.headers]]
        # print(f"table_data = {table_data}")

        # print(f"!!! dataset.dict = {dataset.dict}")
        
        totals_row_style = bold_body_style
        if not is_totals_row:
            totals_row_style = body_style

        # Body rows
        for nn, row in enumerate(dataset.dict):
            # print(f"!!!row = {row}")
            # print(f"!!!row.values() = {row.values()}")
            if nn != len(dataset.dict) - 1:
                table_data.append([Paragraph(clean_cell(cell), body_style) for cell in row.values()])
            else:
                # last row
                table_data.append([Paragraph(clean_cell(cell), totals_row_style) for cell in row.values()])

        num_cols = len(table_data[0])  # number of columns from the first row
        print(f"num_cols = {num_cols}")
        col_width = usable_width / num_cols
        print(f"col_width = {col_width}")
        
        num_rows = len(table_data)
        table = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),  # Use Cyrillic font
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ("BACKGROUND", (0, 0), (-1, 0), lighter_green),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))

        doc.build([table])
        pdf_value = buffer.getvalue()
        buffer.close()
        return pdf_value
    


class XLSX(base_formats.XLSX):
    def export_data(self, dataset, **kwargs):
        xlsx_bytes = super().export_data(dataset, **kwargs)
        stream = BytesIO(xlsx_bytes)
        wb = load_workbook(stream)
        ws = wb.active
        
        col_padding = 4

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        max_est_lines = 1

        for cell in ws[1]:
            value = str(cell.value) if cell.value is not None else ""
            value = _insert_line_breaks(value, words_per_line=2)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

            est_lines = value.count("\n") + 1
            if est_lines > max_est_lines:
                max_est_lines = est_lines

        ws.row_dimensions[1].height = 40 + (max_est_lines - 1) * 8
        
        # --- Set column width based ONLY on header ---
        for cell in ws[1]:
            column_letter = cell.column_letter
            text_line = str(cell.value).split('\n')[0]
            header_length = len(text_line)
            # print(f"text_line = {text_line}")
            ws.column_dimensions[column_letter].width = header_length + col_padding

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell.value = clean_cell(cell.value)
                    text_line = str(cell.value).split('\n')[0]
                    val_len = len(text_line) if cell.value else 0
                    if val_len > max_length:
                        max_length = val_len
                except Exception:
                    pass
            if (max_length + col_padding) > ws.column_dimensions[column].width:
                ws.column_dimensions[column].width = max_length + col_padding

        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    



